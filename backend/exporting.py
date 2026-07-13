"""Level 1 — Excel (.xlsx) export tool (openpyxl).

Deterministic: takes rows (list of dicts) and writes a real spreadsheet under
backend/exports/. Kept separate from the agent code, like reporting.py.
"""

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXPORTS_DIR = Path(__file__).parent / "exports"


def _slug(text: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", (text or "").lower()).strip("-")
    return slug or "export"


def export_rows(title: str, rows: list[dict]) -> dict:
    """Write rows to an .xlsx file and return {path, filename}."""
    EXPORTS_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"{_slug(title)}-{timestamp}.xlsx"
    path = EXPORTS_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Data")[:31]

    headers = list(rows[0].keys()) if rows else ["(no data)"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F6F5C")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([r.get(h) for h in headers])

    # Auto-size columns to their content (clamped).
    for i, h in enumerate(headers, 1):
        longest = len(str(h))
        for r in rows:
            longest = max(longest, len(str(r.get(h, ""))))
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 50)

    ws.freeze_panes = "A2"
    wb.save(path)
    return {"path": str(path), "filename": filename}


if __name__ == "__main__":
    out = export_rows(
        "June Sales",
        [{"month": "June", "amount": 45000}, {"month": "July", "amount": 39800}],
    )
    print("wrote:", out)
