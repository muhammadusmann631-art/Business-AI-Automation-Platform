"""Batch 1 — CSV/Excel import: parse a file, map its columns to a table, and
validate rows before they are inserted.

Deterministic and dependency-light: `csv` for delimited files, `openpyxl`
(already a dep) for .xlsx. No pandas. Parsing/mapping/validation live here;
the actual dedup-aware insert lives in database.import_rows.
"""

import csv
import io
import re

import database as db

MAX_BYTES = 5 * 1024 * 1024  # 5 MB upload cap

# Synonyms: file header (normalised) -> canonical table column. Only columns
# that exist in the target table (database.ADMIN_TABLES) are actually used.
_SYNONYMS: dict[str, list[str]] = {
    "name": ["name", "customer name", "client name", "full name", "product name", "title"],
    "email": ["email", "e-mail", "mail", "email address"],
    "company": ["company", "organization", "organisation", "org", "business"],
    "phone": ["phone", "phone number", "mobile", "contact", "tel", "cell"],
    "city": ["city", "location", "town"],
    "status": ["status", "state"],
    "invoice_number": ["invoice number", "invoice no", "invoice", "inv", "number", "invoice_number"],
    "customer_id": ["customer id", "client id", "customer_id", "cust id"],
    "amount": ["amount", "total", "value", "revenue", "sales", "cost"],
    "due_date": ["due date", "due", "due_date"],
    "paid_date": ["paid date", "payment date", "paid_date"],
    "description": ["description", "desc", "details", "note", "notes", "memo"],
    "category": ["category", "type", "expense category"],
    "price": ["price", "unit price", "rate"],
    "stock": ["stock", "quantity", "qty", "inventory", "units"],
    "month": ["month", "period"],
    "date": ["date", "expense date", "transaction date"],
}

# Columns that must be present (post-mapping) for a row to import.
_REQUIRED: dict[str, list[str]] = {
    "customers": ["name"],
    "invoices": ["invoice_number", "amount"],
    "products": ["name", "price"],
    "sales": ["month", "amount"],
    "expenses": ["category", "amount", "date"],
}
_NUMERIC = {"amount", "price", "stock", "customer_id"}


class ImportError_(ValueError):
    """Bad file / unusable content."""


def _norm(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (h or "").strip().lower()).strip()


def parse_file(filename: str, content: bytes) -> list[dict]:
    """Read a CSV or Excel file into a list of {header: value} dicts."""
    if len(content) > MAX_BYTES:
        raise ImportError_("File too large (max 5 MB).")
    name = (filename or "").lower()

    if name.endswith(".csv") or (not name.endswith((".xlsx", ".xls")) and _looks_text(content)):
        text = _decode(content)
        delimiter = _sniff_delimiter(text)
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        return [{(k or ""): (v or "") for k, v in row.items()} for row in reader]

    if name.endswith((".xlsx", ".xls")):
        return _parse_excel(content)

    raise ImportError_("Unsupported file type. Use .csv or .xlsx.")


def _looks_text(content: bytes) -> bool:
    return b"\x00" not in content[:1024]


def _decode(content: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return content.decode("latin-1", "replace")


def _sniff_delimiter(text: str) -> str:
    sample = text[:2048]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except Exception:
        # fall back to the most common of the usual suspects
        counts = {d: sample.count(d) for d in [",", ";", "\t", "|"]}
        return max(counts, key=counts.get) or ","


def _parse_excel(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    out = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        out.append({headers[i]: ("" if c is None else c) for i, c in enumerate(row) if i < len(headers)})
    wb.close()
    return out


def build_mapping(table: str, headers: list[str]) -> dict[str, str]:
    """Map file headers -> table columns (fuzzy, case-insensitive)."""
    table_cols = db.ADMIN_TABLES.get(table, [])
    mapping: dict[str, str] = {}
    for header in headers:
        n = _norm(header)
        if not n:
            continue
        for col in table_cols:
            if col in mapping.values():
                continue
            candidates = _SYNONYMS.get(col, [col])
            if n == _norm(col) or any(n == _norm(c) for c in candidates):
                mapping[header] = col
                break
    return mapping


def _map_and_validate(table: str, rows: list[dict]):
    """Apply the header mapping + per-cell validation. Returns (clean, errors)."""
    if not rows:
        return [], [], {}
    mapping = build_mapping(table, list(rows[0].keys()))
    required = _REQUIRED.get(table, [])
    missing = [c for c in required if c not in mapping.values()]
    if missing:
        raise ImportError_(
            f"Missing required column(s): {', '.join(missing)}. "
            f"Detected columns: {', '.join(mapping.values()) or 'none'}."
        )

    clean, errors = [], []
    for i, row in enumerate(rows, 1):
        mapped: dict = {}
        bad = None
        for header, col in mapping.items():
            val = row.get(header, "")
            if isinstance(val, str):
                val = val.strip()
            if val in (None, ""):
                continue
            if col in _NUMERIC:
                num = _to_number(val)
                if num is None:
                    bad = f"Row {i}: '{col}' is not a number ({val!r})"
                    break
                val = num
            if col == "email" and "@" not in str(val):
                bad = f"Row {i}: invalid email ({val!r})"
                break
            mapped[col] = val
        if bad:
            errors.append(bad)
            continue
        if any(c not in mapped for c in required):
            errors.append(f"Row {i}: missing required value ({', '.join(required)})")
            continue
        clean.append(mapped)
    return clean, errors, mapping


def _to_number(val):
    try:
        return float(str(val).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def preview(table: str, filename: str, content: bytes) -> dict:
    """Parse + map WITHOUT importing. Returns column mapping + first 5 rows."""
    rows = parse_file(filename, content)
    clean, errors, mapping = _map_and_validate(table, rows)
    return {
        "filename": filename,
        "table": table,
        "total_rows": len(rows),
        "mapping": mapping,
        "columns": sorted(set(mapping.values())),
        "preview": clean[:5],
        "errors": errors[:10],
    }


def do_import(table: str, filename: str, content: bytes) -> dict:
    """Parse, map, validate, and insert (dedup-aware). Returns import result."""
    rows = parse_file(filename, content)
    clean, validation_errors, _mapping = _map_and_validate(table, rows)
    result = db.import_rows(table, clean)
    result["errors"] = (validation_errors + result.get("errors", []))[:20]
    result["filename"] = filename
    return result


if __name__ == "__main__":
    db.seed()
    csv_bytes = (
        "Customer Name;E-mail;Company;City\n"
        "Zainab Imports;zainab@imp.co;Imp Co;Multan\n"
        "Bad Row;not-an-email;X;Y\n"
        "Ahmed Raza;ahmed@techsol.pk;dup;Karachi\n"
    ).encode()
    print("preview:", preview("clients.csv", None) if False else preview("customers", "clients.csv", csv_bytes))
    print("import:", do_import("customers", "clients.csv", csv_bytes))
